from openpyxl import load_workbook
import pyautogui as main
import pyperclip as pc
import re
import time
import pymsgbox
import numpy as np


Excel = "GESTION.xlsx"

Wb = load_workbook(Excel, data_only=True)
sheet = Wb["Hoja1"]

initial_data = [""] * 2


def fila_excel():
    delete = [0,4,9,10]
    for value in sheet.iter_rows(min_row = 1, max_row = 1,values_only=True):
            elementos = list(value)
            for i in sorted(delete, reverse = True):
                del elementos[i]

    elementos[1] = elementos[1].strftime("%Y%m%d")
    elementos[3] = elementos[3].strftime("%Y%m%d")
    elementos = initial_data + elementos

    return elementos 

